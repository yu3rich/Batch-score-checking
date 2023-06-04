from openpyxl import load_workbook


# 读取excel信息模块
class Lookup:
    def __init__(self, file_xls):
        self.file_xls = file_xls
        wb = load_workbook(self.file_xls, read_only=True, data_only=False)
        self.sheet = wb.active

    # 单一单元格获取模式
    def data(self, row, col):
        # 读取账号密码信息
        datas = self.sheet.cell(row=row, column=col).value
        return datas

    # 逐行获取模式
    def data_ss(self, row, col):
        ob = []
        a = 0
        while True:
            datas = self.sheet.cell(row=row+a, column=col).value
            a = a+1
            if datas is None:
                return ob
            else:
                ob.append(datas)
